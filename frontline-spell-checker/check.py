import re
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from spellchecker import SpellChecker

def remove_html_tags(text):
    soup = BeautifulSoup(text, "html.parser")
    return soup.get_text()

def remove_adjacent_digits(text):
    pattern = r'(?<=\D)\d+|\d+(?=\D)'
    replaced_text = re.sub(pattern, '', text)
    return replaced_text


def check_spelling(text):
    spell = SpellChecker()
    words = text.split()
    misspelled = spell.unknown(words)
    return misspelled


def main():
    filename = "ListOfFieldsTXDev.xlsx"
    wb = load_workbook(filename=filename)
    ws = wb.active
    i = 0
    for row in ws.iter_rows(min_row=2):
        if i == 1000: break
        if not row[2].value:
            continue

        temp_row = remove_html_tags(row[2].value)
        temp_row = temp_row.lower()
        characters_to_replace = [",", ".", ":", "?", "(", ")", "-", "/", ";", "*", "[", "]", "{", "}", "..."]
        for char in characters_to_replace:
            temp_row = temp_row.replace(char, " ")
        temp_row = remove_adjacent_digits(temp_row)

        corrected_value = check_spelling(temp_row)
        corrected_value = ', '.join(corrected_value)
        print(temp_row, corrected_value)
        row[1].value = corrected_value
        i += 1
    wb.save(filename)


if __name__ == "__main__":
    main()