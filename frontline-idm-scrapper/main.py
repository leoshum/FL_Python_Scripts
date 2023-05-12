import sys
import time
import validators
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

def login_user(driver):
	username_field = driver.find_element(By.CSS_SELECTOR, "kendo-textbox[formcontrolname='username'] input")
	password_field = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
	submit_btn = driver.find_element(By.CSS_SELECTOR, "form button")
	username_field.send_keys("SupportAdmin")
	password_field.send_keys("tkztNANK3dLxD2XyJt")
	submit_btn.click()
	

def open_website(driver, url, user):
	username_field = driver.find_element(By.CSS_SELECTOR, "kendo-textbox[formcontrolname='username'] input")
	host_field = driver.find_element(By.CSS_SELECTOR, "kendo-textbox[formcontrolname='host'] input")
	navigate_button = driver.find_element(By.CSS_SELECTOR, ".k-buttons-end button")
	host_field.clear()
	username_field.clear()
	username_field.send_keys(user)
	host_field.send_keys(url)
	navigate_button.click()


def get_website_version(driver, tab):
	driver.switch_to.window(tab)
	return WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#plcContent_ctlPageFooter_lblVersion, #lblVersion"))).text.replace("Version ", "")


def main():
    options = Options()
    options.headless = True
    driver = webdriver.Chrome(options=options)
    driver.get("https://support-tech.acceliplan.com/login")
    login_user(driver)
    time.sleep(3)
    wb = load_workbook(sys.argv[1], data_only=True)
    wb_sheet = wb.active
    for row in wb_sheet.iter_rows(min_row=3):
          if row[1].value != None and validators.url(row[1].value) and row[16].value == "Active":
            print(f"Processing: {row[1].value}")
            parent = driver.window_handles[0]
            open_website(driver, row[1].value, "PMGMT")
            time.sleep(3)
            build_version = ""
            try:
                build_version = get_website_version(driver, driver.window_handles[1])
            except:
                print(f"Error in processing {row[1].value}")
                build_version = ""
            row[4].value = build_version
            driver.close()
            driver.switch_to.window(parent)
    wb.save(sys.argv[1])

    
if __name__ == "__main__":
	main()