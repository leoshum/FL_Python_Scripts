from datetime import datetime, timedelta
import json
from os import environ, path
import os
import re
import subprocess
import aiohttp
import asyncio
import types
import pytz

import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd

def configure():
    token_name = 'TEAM_CITY_TOKEN'
    teamcity_token = environ.get(token_name)
    if teamcity_token is None:
        print(f"TeamCity access token not found in environments variable with name [{token_name}]!")
        exit()
    
    token_name = 'JIRA_TOKEN'
    jira_token = environ.get(token_name)
    if jira_token is None:
        print(f"Jira access token not found in environments variable with name [{token_name}]!")
        exit()

    root_directory = path.dirname(__file__)
    configuration_path = f"{root_directory}\\configuration.json"
    if not path.exists(configuration_path):
        print(msg=f"Configuration file not found in [{configuration_path}] path!")
        exit()
    with open(configuration_path) as configuration_file:
        configuration = json.loads(configuration_file.read())
    
    configuration['jira_url'] = f'https://{configuration.get("jira_host")}/rest/api/2/search'
    configuration['teamcity_url'] = f'https://{configuration.get("teamcity_host")}/app/rest'
    
    configuration['jira_token'] = jira_token
    configuration['teamcity_token'] = teamcity_token

    configuration['find_script_path'] = f"{root_directory}\\find_tickets.ps1"
    configuration['switch_script_path'] = f"{root_directory}\\switch_and_update_branch.ps1"
    configuration['result_path'] = f"{root_directory}\\result.xlsx"
    configuration['result_json_path'] = f"{root_directory}\\result.json"
    configuration['builds_path'] = f"{root_directory}\\{configuration.get('builds_file')}"
    

    print('Application configured')
    return configuration

async def get_issues_from_jira(session, jira_url, login, jira_token, stored_issues):
    # Authenticate with Jira
    auth = aiohttp.BasicAuth(login=login, password=jira_token)
    
    jql_query = 'labels=jira_escalated and project="CW-0575 (Accelify)" and status not in ("waiting for response", open, DevReady, "in progress")'
    start_at = 0
    max_results = 50
    headers = {'Accept': 'application/json'}
    numbers = [issue.get('number') for issue in stored_issues]
    date_format = "%Y-%m-%dT%H:%M:%S.%f%z"
    # Calculate the date six months ago
    end_date = datetime.now(pytz.utc) - timedelta(days=30*4)
    total_issues = 1
    # result = []
    # number_regex = re.compile(r"^CW0575-(\d+)$")
    # last_number = re.findall(number_regex, stored_issues[0].get('number'))[0]

    # Make the initial API call to get the total number of issues
    # url = f'{jira_url}?jql={jql_query}&startAt={start_at}&maxResults={max_results}'
    # async with session.get(url, headers=headers, auth=auth) as response:
    #     if response.status == 200:
    #         data = await response.json()
    #         total_issues = data['total']
    #         result.extend(data['issues'])
    #         print(f'Get [{start_at}] - [{start_at + max_results - 1}] issues.')
    #         start_at += max_results
    #     else:
    #         print(f'Request failed with status {response.status}: {response.reason}')
    #         return
    # Loop through the paginated API responses
    while start_at < total_issues:# and start_at < 50
        # Set up the API request with the appropriate startAt and maxResults values
        url = f'{jira_url}?jql={jql_query}&startAt={start_at}&maxResults={max_results}'

        # Make the API call with the authenticated session
        async with session.get(url, headers=headers, auth=auth) as response:
            if response.status == 200:
                data = await response.json()
                if total_issues == 1 and data.get('total') != 1:
                    total_issues = data['total']

                # result.extend(data['issues'])
                print(f'Get [{start_at}] - [{start_at + max_results - 1}] issues from Jira.')
                for issue in data.get('issues'):
                    if issue.get('key') in numbers:
                        if datetime.strptime(issue.get('fields').get('updated'), date_format) < end_date: return
                        continue
                    # if last_number >= re.findall(number_regex, issue.get('key'))[0]: return
                    yield issue                 
            else:
                print(f'Request failed with status {response.status}: {response.reason}')
                return

        # Increment the startAt value to get the next page of results
        start_at += max_results
        
    # with open(result_path, mode='w', encoding='UTF-8') as file:
    #     file.write(json.dumps([issue['fields']['status']['name'] for issue in result], indent=2, ensure_ascii=False))
    # return [{ 'key': issue['key'], 'status' : issue['fields']['status']['name']} for issue in result]

async def configure_git(script_path, repository_path, branch_name, key_path):
    process = await asyncio.create_subprocess_exec(
        'powershell', '-File', script_path, '-Repository_path', repository_path, '-Branch_name', branch_name, '-Key_Path', key_path,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE
    )

    # read the stdout and stderr asynchronously
    stdout, stderr = await process.communicate()
    print(stdout.decode(), stderr.decode())

async def find_in_git(ticket, script_path, repository_path):
    process = await asyncio.create_subprocess_exec(
        'powershell', '-File', script_path, '-Repository_path', repository_path, '-Ticket_number', ticket['key'],
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE
    )
    stdout, stderr = await process.communicate()

    if process.returncode != 0:
        print(f'Ticket [{ticket["key"]}] not found in Git')
        return None
    
    print(f'Ticket [{ticket["key"]}] found in Git')
    return json.loads(stdout.decode('utf-8'))[0]
    # task_result = subprocess.run(['powershell', '-File', script_path, '-Repository_path', repository_path, '-Ticket_number', ticket['key']], capture_output=True)
    # if task_result.returncode != 0:
    #     result.append({
    #         'number' : ticket['key'],
    #         'status' : ticket['status'],
    #         'commits' : []
    #     })
    #     not_founded.append({                
    #         'number' : ticket['key']
    #     })
    #     print(f'Ticket [{ticket["key"]}] not founded')
    # else:
    #     output = json.loads(task_result.stdout.decode('utf-8'))
    #     result.append({
    #         'number' : ticket['key'],
    #         'status' : ticket['status'],
    #         'commits' : output
    #     })
    #     print(f'Ticket [{ticket["key"]}] founded')

    # def find_in_git(tickets, find_script_path, switch_script_path, repository_path, branch_name, key_path):
    #     result = []
    #     not_founded = []
    #     task_result = subprocess.run(['powershell', '-File', switch_script_path, '-Repository_path', repository_path, '-Branch_name', branch_name, '-Key_Path', key_path])
    #     if task_result.returncode != 0:
    #         output = json.loads(task_result.stdout.decode('utf-8'))
    #         raise f'Can not switch branch [{output}]'

    #     for ticket in tickets:
    #         task_result = subprocess.run(['powershell', '-File', find_script_path, '-Repository_path', repository_path, '-Ticket_number', ticket['key']], capture_output=True)
    #         if task_result.returncode != 0:
    #             result.append({
    #                 'number' : ticket['key'],
    #                 'status' : ticket['status'],
    #                 'commits' : []
    #             })
    #             not_founded.append({                
    #                 'number' : ticket['key']
    #             })
    #             print(f'Ticket [{ticket["key"]}] not founded')
    #         else:
    #             output = json.loads(task_result.stdout.decode('utf-8'))
    #             result.append({
    #                 'number' : ticket['key'],
    #                 'status' : ticket['status'],
    #                 'commits' : output
    #             })
    #             print(f'Ticket [{ticket["key"]}] founded')
    #     return result

def parse_commit(commit):
    commit_regex = re.compile(r"^commit\s+(\w+)")
    mathces = re.findall(commit_regex, commit[0])
    return mathces[0]

def parse_tickets(tickets):
    commit_regex = re.compile(r"^commit\s+(\w+)") #'^commit [0-9a-f]{40}$'
    author_regex = re.compile(r"^Author:\s+(.*)\s+<.*>")
    email_regex = re.compile(r"^Author:\s+.*\s+<(.*)>")
    date_regex = re.compile(r"^Date:\s+(.*)")
    for ticket in tickets:
        if not ticket.get('commits'): continue
        for index, commit in enumerate(ticket.get('commits')):
            parsed_commit = {
                'sha' : None,
                'author' : None,
                'email' : None,
                'date' : None,
                'description' : []
            }
            for line in commit:
                if parse_line(parsed_commit, 'sha', line, commit_regex): continue
                if parse_line(parsed_commit, 'author', line, author_regex) and parse_line(parsed_commit, 'email', line, email_regex): continue
                if parse_line(parsed_commit, 'date', line, date_regex):
                    parsed_commit['date'] = datetime.strptime(parsed_commit.get('date'), "%a %b %d %H:%M:%S %Y %z").isoformat(sep=' ', timespec='seconds')
                    continue
                parsed_commit['description'].append(line)
            parsed_commit['description'] = '\r\n'.join(parsed_commit.get('description'))
            ticket['commits'][index] = parsed_commit

    # for ticket in tickets:
    #     for commit in ticket.get('commits'):
    #         parsed_commit = {
    #             'sha' : None,
    #             'author' : None,
    #             'email' : None,
    #             'date' : None,
    #             'description' : []
    #         }
    #         for line in commit:
    #             if parse_line(parsed_commit, 'sha', line, commit_regex): continue
    #             if parse_line(parsed_commit, 'author', line, author_regex) and parse_line(parsed_commit, 'email', line, email_regex): continue
    #             if parse_line(parsed_commit, 'date', line, date_regex):
    #                 parsed_commit['date'] = datetime.datetime.strptime(parsed_commit.get('date'), "%a %b %d %H:%M:%S %Y %z")
    #                 continue
    #             parsed_commit.get('description').append(line)
    #         parsed_commit['description'] = '\r'.join(parsed_commit.get('description'))
    #         ticket.get('commits')[ticket.get('commits').index(commit)] = parsed_commit


                # if not parsed_commit.get('sha'):
                #     commit_sha = re.findall(commit_regex, line)
                #     if len(commit_sha) != 0:
                #         parsed_commit['sha'] = commit_sha[0]
                #         continue
                # if 

                # author_and_email = re.findall(author_regex, line)[0]
                # author = author_and_email[0]
                # email = author_and_email[1]

                # date_string = re.findall(date_regex, line)[0]
                # date = datetime.datetime.strptime(date_string, "%a %b %d %H:%M:%S %Y %z")

                # print('')

def parse_line(commit, key, line, regex):
    if not commit.get(key):
         
            return True
    return False

async def get_builds(session, teamcity_url, teamcity_token, build_id, builds_path):
    if os.path.exists(builds_path):
        with open(builds_path,mode='r',encoding='UTF-8') as file:
            stored = json.load(file)
    else: stored = None

    headers = {
        'Accept': 'application/json',
        'Authorization' : f'Bearer {teamcity_token}'
    }
    builds = []
    start = 0
    count = 100
    stop = False
    team_city_time_format = '%Y%m%dT%H%M%S%z'

    while not stop:
        url = f'{teamcity_url}/builds?locator=defaultFilter:false,buildType:{build_id},count:{count},start:{start},status:SUCCESS,running:false'
        async with session.get(url, headers=headers) as response:
            if response.status == 200:
                data = await response.json()
                if data.get('count') == 0: 
                    stop = True
                    break
                for build in data.get('build'):
                    if stored and len(stored) > 0 and datetime.strptime(build.get('finishOnAgentDate'), team_city_time_format) <= datetime.strptime(stored[0].get('finishOnAgentDate'), team_city_time_format):
                        builds.extend(stored)
                        stop = True
                        break
                    builds.append(build)
                    change_start = 0
                    build['changes'] = []
                    while True:
                        url = f"{teamcity_url}/changes?build={build['id']}&fields=change(version,date)&start={change_start}&count:{count}"
                        async with session.get(url, headers=headers) as response:
                            if response.status == 200:
                                data = await response.json()
                                build['changes'].extend(data.get('change'))
                                if len(data.get('change')) < 100 or datetime.strptime(data.get('change')[0].get('date'), team_city_time_format).year < 2020:
                                    break
                                # for change in data['change']:
                                #     changes[change['version']] = build['number']
                        change_start += count
        start += count
    
    with open(builds_path,mode='w',encoding='UTF-8') as file:
        file.write(json.dumps(builds,indent=2,ensure_ascii=False))
    result = { change['version']: build['number'] for build in builds for change in build['changes'] }
    print(f'[{len(result)}] Builds received from the TeamCity')
    return result

async def find_build(session, tickets, teamcity_url, teamcity_token, build_id):
    headers = {
        'Accept': 'application/json',
        'Authorization' : f'Bearer {teamcity_token}'
    }

    changes = {}
    url = f'{teamcity_url}/builds?locator=defaultFilter:false,buildType:{build_id}'
    async with session.get(url, headers=headers) as response:
                    if response.status == 200:
                        data = await response.json()
                        for build in data['build']:
                            url = f"{teamcity_url}/changes?locator=build:(id:{build['id']})&fields=change(version)"
                            async with session.get(url, headers=headers) as response:
                                if response.status == 200:
                                    data = await response.json()
                                    for change in data['change']:
                                        changes[change['version']] = build['number']
                        #     url = f'{teamcity_url}/builds?locator=defaultFilter:false,buildType:{build_id}'
                        # builds = data.get('build')
                        # if builds:
                        #     commit['build'] = []
                        #     for build in builds:
                        #         commit['build'].append(build.get('number'))
                        # else:
                        #     commit['build'] = None
                    else:
                        print(f'Request failed with status {response.status}: {response.reason}')

    


    for ticket in tickets:
        if not ticket.get('commits'): continue
        for commit in ticket.get('commits'):
            sha = commit.get('sha')
            commit['build'] = changes.get(sha)
            #https://teams.acceliplan.com/app/rest/builds/?locator=buildType:BuildNewArchitecture_Main_BuildVersion,changes:change:(version:81f74c1a923f4cfc00f27dd1abeca87a85e86ce7)
            #https://teams.acceliplan.com/httpAuth/app/rest/builds/?locator=buildType:BuildNewArchitecture_Main_BuildVersion,revision:81f74c1a923f4cfc00f27dd1abeca87a85e86ce7
            # url = f'{teamcity_url}/changes?locator=build:(id:{build_id})'
            
            # url = f'{teamcity_url}/changes?locator=?locator=buildType:{build_id},version:{sha}'
            # async with session.get(url, headers=headers) as response:
            #     if response.status == 200:
            #         data = await response.json()
            #         builds = data.get('build')
            #         if builds:
            #             commit['build'] = []
            #             for build in builds:
            #                 commit['build'].append(build.get('number'))
            #         else:
            #             commit['build'] = None
            #     else:
            #         print(f'Request failed with status {response.status}: {response.reason}')

def write_to_excel(tickets, path):
                                            
                                            #add status ticket!!!!!!!!!!!!!!!!!

    # Load Excel file

    # Load Excel file
    # filename = 'output.xlsx'
    # book = load_workbook(filename)

    # # Load sheet
    # sheet_name = 'Sheet1'
    # df = pd.read_excel(filename, sheet_name)

    # # Write DataFrame to sheet
    # with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    #     writer.book = book
    #     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #     df.to_excel(writer, sheet_name=sheet_name, index=False)

    #     # Adjust column width and row height based on content
    #     for column in writer.sheets[sheet_name].columns:
    #         max_length = 0
    #         column_name = column[0].column_letter  # Get the column name
    #         for cell in column:
    #             try:
    #                 if len(str(cell.value)) > max_length:
    #                     max_length = len(str(cell.value))
    #             except:
    #                 pass
    #         adjusted_width = (max_length + 2) * 1.2
    #         writer.sheets[sheet_name].column_dimensions[column_name].width = adjusted_width
            
    #     for row in writer.sheets[sheet_name].rows:
    #         max_height = 0
    #         for cell in row:
    #             try:
    #                 cell_value = str(cell.value)
    #                 line_count = cell_value.count('\n') + 1
    #                 if line_count > max_height:
    #                     max_height = line_count
    #             except:
    #                 pass
    #         adjusted_height = max_height * 15
    #         writer.sheets[sheet_name].row_dimensions[row[0].row].height = adjusted_height

    df = pd.DataFrame(tickets)

    # write the DataFrame to an Excel file
    df.to_excel(path, index=False)
    # df = pd.json_normalize(tickets, record_path=['commits'], meta=['number', 'status'])
    # df.to_excel(path, index=False)

async def main():
    tickets = []
    configuration = configure()
    with open(configuration.get('result_json_path'), mode='r', encoding='UTF-8') as file:
        stored_issues = json.load(file)
    stored_issues = stored_issues[100:len(stored_issues)]
    async with aiohttp.ClientSession() as jira_session:
        issues = get_issues_from_jira(jira_session, configuration.get('jira_url'), configuration.get('jira_login'), configuration.get('jira_token'), stored_issues)
        # with open("C:\\Users\\mykha\\source\\repos\\fl_python\\FL_Python_Scripts\\frontline-ticket-parser\\tickets.json", mode='r',encoding='UTF-8') as file:
        #     issues = json.load(file)
        # for index, ticket in enumerate(issues):
        #     issues[index] = {
        #         "key" : ticket.replace(" ", ""),
        #         'fields' : { 'status' : {'name':'...'}}
        #     }
        await configure_git(configuration.get('switch_script_path'), configuration.get('repository_path'), configuration.get('git_branch'), configuration.get('ssh_key_path'))
        
        async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=False),trust_env=True) as teamcity_session:
            builds = get_builds(teamcity_session, configuration.get('teamcity_url'), configuration.get('teamcity_token'), configuration.get('teamCity_build_id'), configuration.get('builds_path'))
            # builds = await builds
            async for ticket in issues:
                output = await find_in_git(ticket, configuration.get('find_script_path'), configuration.get('repository_path'))
                if not output:
                    tickets.append({
                        'number' : f'https://frontlinetechnologies.atlassian.net/browse/{ticket.get("key")}',
                        'status' : ticket['fields']['status']['name'],
                        'version' : None
                    })
                    continue
                sha = parse_commit(output)
                if type(builds) is types.CoroutineType:
                    builds = await builds
                tickets.append({
                    'number' : ticket['key'],
                    'status' : ticket['fields']['status']['name'],
                    'version' :  builds.get(sha)
                })
    tickets.extend(stored_issues)
    write_to_excel(tickets, configuration.get('result_path'),)
    with open(configuration.get('result_json_path'), mode='w', encoding='UTF-8') as file:
        file.write(json.dumps(tickets, indent=2, ensure_ascii=False))


    #     print(f'Founded [{len(issues)}] tickets in Jira')
    #     result = find_in_git(issues, configuration.get('find_script_path'), configuration.get('switch_script_path'), configuration.get('repository_path'), configuration.get('git_branch'), configuration.get('ssh_key_path'))
    #     parse_tickets(result)
    # async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=False),trust_env=True) as jira_session: #verify_
        
    #     await find_build(jira_session, result, configuration.get('teamcity_url'), configuration.get('teamcity_token'), configuration.get('teamCity_build_id'))
    #     write_to_excel(result, configuration.get('result_path'),)
    #     with open(configuration.get('result_json_path'), mode='w', encoding='UTF-8') as file:
    #         file.write(json.dumps(result, indent=2, ensure_ascii=False))


# Run the script
asyncio.run(main())


    # build_id = configuration.get('teamCity_build_id') #'BuildNewArchitecture_Main_BuildVersion'
    # login = configuration.get('jira_login') #'mtolstikhin@frontlineed.com'#lshum
    # ssh_key_path = configuration.get('ssh_key_path') #'~/.ssh/a2c'
    # branch_name = configuration.get('git_branch') #'main'

    # #repository_path = configuration.get('repository_path') #'C:\\Users\\mykha\\source\\repos\\CW-0575-IEP'        
    # #root_directory = path.dirname(__file__)
    # find_script_path = f"{root_directory}\\find_tickets.ps1"
    # switch_script_path = f"{root_directory}\\switch_and_update_branch.ps1"
    # result_path = f"{root_directory}\\result.json"

    #JIRA query: you can try this search in zendesk - jira_escalated brand:plan status:new status:hold status:open status:pending (edited) 