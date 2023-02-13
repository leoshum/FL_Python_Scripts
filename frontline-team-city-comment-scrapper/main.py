from asyncio import get_event_loop
import json
import logging
from os import environ, path
from sys import exit,argv
from datetime import datetime
import aiohttp

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from packaging import version


logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

token_name = 'TEAM_CITY_TOKEN'
token = environ.get(token_name)
if token is None:
    logging.error(msg=f"TeamCity bearer token not found in environments variable with name [{token_name}]!")
    exit()

ticket_match = "https://frontlinetechnologies.atlassian.net/browse/CW0575-"
ticket_number_length = 5

root_directory = path.dirname(argv[0])
configuration_path = f".{root_directory}\\configuration.json"
if not path.exists(configuration_path):
    logging.error(msg=f"Configuration file not found in [{configuration_path}] path!")
    exit()

with open(configuration_path) as configuration_file:
    configuration = json.loads(configuration_file.read())

version_key = 'version'
if version_key not in configuration:
    logging.error(msg=f"[{version_key}] field does not exist in configuration file!")
    exit()

first_version_key = 'first'
if first_version_key not in configuration[version_key]:
    logging.error(msg=f"[{first_version_key}] field does not exist in [{version_key}] field in configuration file!")
    exit()
from_version = version.parse(configuration[version_key][first_version_key])

last_version_key = 'last'
if last_version_key not in configuration[version_key]:
    logging.error(msg=f"[{last_version_key}] field does not exist in [{version_key}] field in configuration file!")
    exit()
to_version = version.parse(configuration[version_key][last_version_key])
if to_version < from_version:
    temp = to_version
    to_version = from_version
    from_version = to_version

comments_path = f".{root_directory}\\{from_version.__str__()}-{to_version.__str__()}.xlsx"

host_key = 'host'
if host_key not in configuration:
    logging.error(msg=f"[{host_key}] field does not exist in configuration file!")
    exit()
baseurl = f"https://{configuration[host_key]}/app/rest"

branch = argv[1] if len(argv) > 1 else configuration['branch']
if branch == 'prod' or branch == 'production' or branch == 'main':
    branch = 'Production'
    job_id = 'BuildNewArchitecture_Main_BuildVersion'
elif branch == 'release':
    branch = 'Release'
    job_id = 'BuildNewArchitecture_TagFix_BuildVersion'
elif branch == 'develop' or branch == 'dev':
    branch = 'Develop'
    job_id = 'BuildNewArchitecture_Trunk_BuildVersion'
else:
    logging.error(msg=f"[{branch}] name of branch does not much to any cases!")
    exit()

columns = ['Version','Build Date','Author','Commit Date','Jira Ticket','Comment']
width = {
    columns[0] : 0,
    columns[1] : 0,
    columns[2] : 0,
    columns[3] : 0,
    columns[4] : 0,
    columns[5] : 0
}

book = Workbook()
sheet = book.active
sheet.title = "Comments"
sheet.append([branch])
sheet.append(columns)

async def main():
    start = 0
    current_version = from_version
    async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(verify_ssl=False),trust_env=True) as session:
        need_to_parse = True
        while need_to_parse:

            url = f"{baseurl}/builds?locator=defaultFilter:false,start:{start},buildType:{job_id}"
            list_builds = await get(url, token, session)
            if len(list_builds['build']) == 0:
                logger.info(f"Recived empty builds list.")
                need_to_parse = False
                break
            logger.info(f"Detect [{len(list_builds['build'])}] builds for [{job_id}] job, started from [{start}].")
            
            for build in list_builds['build']:
                current_version = version.parse(build['number'])
                if current_version < from_version or current_version > to_version: continue
                
                url = f"{baseurl}/changes?locator=build:(id:{build['id']})&fields=change(username,date,comment)"
                changes = await get(url, token, session)
                logger.info(f"Recived details for [{build['number']}] build.")

                updateLength(build['number'], width, 'Version')
                build_date = str(datetime.strptime(build['finishOnAgentDate'][0:15], '%Y%m%dT%H%M%S'))
                updateLength(build_date, width, 'Build Date')
                
                for change in changes['change']:
                    
                    change_date = str(datetime.strptime(change['date'][0:15], '%Y%m%dT%H%M%S'))

                    change['comment'] = change['comment'].replace('\n', '')
                    if ticket_match in change['comment']:
                        ticket = change['comment'][change['comment'].index(ticket_match):len(ticket_match)+ticket_number_length]
                        comment = change['comment'][len(ticket_match)+ticket_number_length:]
                    else:
                        ticket = ''
                        comment = change['comment']


                    updateLength(change['username'], width, 'Author')
                    updateLength(change_date, width, 'Commit Date')
                    updateLength(ticket, width, 'Jira Ticket')
                    updateLength(comment, width, 'Comment')
                    sheet.append([build['number'], build_date, change['username'], change_date, ticket, comment])
            start += 100

        for i, column_width in enumerate(width.values(),1):  # ,1 to start at 1
            sheet.column_dimensions[get_column_letter(i)].width = column_width + 3

        book.save(filename=comments_path)

async def get(url, token, session):
    headers = {
        "Authorization" : f"Bearer {token}",
        "Accept" : "application/json"}

    async with session.get(url = url, headers = headers) as response:
        if response.status != 200:
            raise Exception("Error in API call: " + await response.text())
        return await response.json()

def updateLength(value, width, name):
    length = len(value)
    if length > width[name]:
        width[name] = length

get_event_loop().run_until_complete(main())