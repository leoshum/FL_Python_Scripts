import json
import logging
from asyncio import create_task, gather, get_event_loop
from datetime import datetime
from os import environ, path
import re
from sys import argv, exit

import aiohttp
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from packaging import version

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

token_name = 'TEAM_CITY_TOKEN'
token = environ.get(token_name)
if token is None:
    logging.error(msg=f"TeamCity bearer token not found in environments variable with name [{token_name}]!")
    exit()

ticket_match = "https://frontlinetechnologies.atlassian.net/browse/CW0575-"
ticket_number_length = 5

root_directory = path.dirname(__file__)
configuration_path = f"{root_directory}\\configuration.json"
if not path.exists(configuration_path):
    logging.error(msg=f"Configuration file not found in [{configuration_path}] path!")
    exit()

with open(configuration_path) as configuration_file:
    configuration = json.loads(configuration_file.read())

version_key = 'version'
if version_key not in configuration:
    logging.error(msg=f"[{version_key}] field does not exist in configuration file!")
    exit()

branch = argv[1] if len(argv) > 1 else configuration['branch']
if branch == 'prod' or branch == 'Production' or branch == 'main':
    branch = 'Production'
    job_id = 'BuildNewArchitecture_Main_BuildVersion'
elif branch == 'Release':
    branch = 'Release'
    job_id = 'BuildNewArchitecture_TagFix_BuildVersion'
elif branch == 'Develop' or branch == 'dev':
    branch = 'Develop'
    job_id = 'BuildNewArchitecture_Trunk_BuildVersion'
else:
    logging.error(msg=f"[{branch}] name of branch does not much to any cases!")
    exit()

first_version_key = 'first'
from_version = version.parse(configuration[version_key][branch][first_version_key])

last_version_key = 'last'
to_version = version.parse(configuration[version_key][branch][last_version_key])
if to_version < from_version:
    temp = to_version
    to_version = from_version
    from_version = to_version

release_dates = {}
for version_name in configuration[version_key]['release_date']:
    release_dates[version_name] = datetime.strptime(configuration[version_key]['release_date'][version_name], '%Y-%m-%dT%H:%M:%SZ')


comments_path = f"{root_directory}\\{from_version.__str__()}-{to_version.__str__()}.xlsx"

host_key = 'host'
if host_key not in configuration:
    logging.error(msg=f"[{host_key}] field does not exist in configuration file!")
    exit()
baseurl = f"https://{configuration[host_key]}/app/rest"

columns = ['Version','Build Date','Author','Commit Date','Jira Ticket','Comment']
width = {
    columns[0] : 0,
    columns[1] : 0,
    columns[2] : 0,
    columns[3] : 0,
    columns[4] : 0,
    columns[5] : 0
}

book = Workbook()
sheet = book.active
sheet.title = "Comments"
sheet.append([branch])
sheet.append(columns)

async def main():
    start = 0
    current_version = from_version
    builds = []
    async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(ssl=False),trust_env=True) as session:
        need_to_parse = True
        tasks = []
        while need_to_parse:
            url = f"{baseurl}/builds?locator=defaultFilter:false,start:{start},buildType:{job_id}"
            list_builds = await get(url, token, session)
            if len(list_builds['build']) == 0:
                logger.info(f"Recived empty builds list.")
                need_to_parse = False
                break
            logger.info(f"Detect [{len(list_builds['build'])}] builds for [{job_id}] job, started from [{start}].")
            
            for build in list_builds['build']:
                number = build.get('number')
                finish_date = build.get('finishOnAgentDate')
                if not number or not finish_date: continue
                try:
                    current_version = version.parse(number)
                except:
                    continue
                if current_version < from_version or current_version > to_version: continue
                
                tasks.append(create_task(get_build_info(builds, build, release_dates.get('.'.join([str(current_version.major), str(current_version.minor)])), session)))
                
            start += 100

        await gather(*tasks)
    builds.sort(key=lambda build: datetime.strptime(build[3], '%Y-%m-%d %H:%M:%S'))
    builds.sort(key=lambda build: version.parse(build[0]))
    for build in builds:
        sheet.append(build)
    
    for i, column_width in enumerate(width.values(),1):  # ,1 to start at 1
        sheet.column_dimensions[get_column_letter(i)].width = column_width + 3

    book.save(filename=comments_path)

pattern = r'(https://frontlinetechnologies.atlassian.net/browse/CW0575-\d{5})(.*?)(?=(https://frontlinetechnologies.atlassian.net/browse/CW0575-\d{5})|$)'
async def get_build_info(builds, build, date, session):
    changes_start = 0
    changes_count = 100
    is_continue = True
    while is_continue:
        url = f"{baseurl}/changes?locator=count:{changes_count},start:{changes_start},build:(id:{build['id']})&fields=change(username,date,comment)"
        changes = await get(url, token, session)
        changes_length = len(changes['change'])
        if changes_length == 0: break
        logger.info(f"Recived details for [{build['number']}] build.")

        updateLength(build['number'], width, 'Version')
        build_date = str(datetime.strptime(build['finishOnAgentDate'][0:15], '%Y%m%dT%H%M%S'))
        updateLength(build_date, width, 'Build Date')
        
        for change in changes['change']:
            
            change_date = datetime.strptime(change['date'][0:15], '%Y%m%dT%H%M%S')
            if date and change_date < date:
                is_continue = False
                print("End date")
                break
            change['comment'] = change['comment'].replace('\n', '')

            matches = re.findall(pattern, change.get('comment'))

            for match in matches:
                ticket = match[0]
                comment = match[1]
                updateLength(change['username'], width, 'Author')
                updateLength(str(change_date), width, 'Commit Date')
                updateLength(ticket, width, 'Jira Ticket')
                updateLength(comment, width, 'Comment')
                builds.append([build['number'], build_date, change['username'], str(change_date), ticket, comment])
            if len(matches) == 0:
                ticket = ''
                comment = change['comment']
                updateLength(change['username'], width, 'Author')
                updateLength(str(change_date), width, 'Commit Date')
                updateLength(ticket, width, 'Jira Ticket')
                updateLength(comment, width, 'Comment')
                builds.append([build['number'], build_date, change['username'], str(change_date), ticket, comment])
        
        if changes_length < changes_count: break
        changes_start += changes_count

async def get(url, token, session):
    headers = {
        "Authorization" : f"Bearer {token}",
        "Accept" : "application/json"}

    async with session.get(url = url, headers = headers) as response:
        if response.status != 200:
            raise Exception("Error in API call: " + await response.text())
        return await response.json()

def updateLength(value, width, name):
    length = len(value)
    if length > width[name]:
        width[name] = length

#Add scrap additional changes 
#end by date 15 marta 

get_event_loop().run_until_complete(main())