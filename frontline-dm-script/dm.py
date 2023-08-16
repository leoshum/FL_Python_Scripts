import os
import sys
import time
import logging
import argparse
import validators
from datetime import datetime
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException, NoSuchElementException, StaleElementReferenceException

package_path = os.path.abspath('..')
sys.path.append(package_path)
from frontline_selenium.selenium_helper import SeleniumHelper
from frontline_selenium.support_tech_helper import SupportTech

class HideBacktraceFormatter(logging.Formatter):
    def formatException(self, exc_info):
        tb = super().formatException(exc_info)
        return HideBacktraceFormatter.removeBackTrace(tb)
    
    def format(self, record: logging.LogRecord):
        record.msg = HideBacktraceFormatter.removeBackTrace(record.msg)
        return super().format(record)
    
    @staticmethod
    def removeBackTrace(record):
        tb_lines = str(record).splitlines()
        filtered_tb_lines = []
        skip_slce = False
        for line in tb_lines:
            if "Backtrace:" in line:
                skip_slce = True
            if "Traceback (most recent call last):" in line:
                skip_slce = False
            if not skip_slce:
                filtered_tb_lines.append(line)
        return "\n".join(filtered_tb_lines)

def split_path(path):
    folders = []
    head, tail = os.path.split(path)
    
    while tail:
        folders.insert(0, tail)
        head, tail = os.path.split(head)
    return folders


def configure_logger(file_name: str, processing_filename: str) -> logging.Logger:
    logger = logging.getLogger("main")
    logger.setLevel(logging.DEBUG)

    formatter = HideBacktraceFormatter("%(asctime)s - %(message)s", datefmt="%m-%d-%y_%H:%M")
    timestamp = datetime.now().strftime("%m-%d-%y_%H-%M")
    parts = split_path(processing_filename)
    folders = parts[0:len(parts)-1]
    filename = parts[-1]
    fh = logging.FileHandler(f"{file_name}_{'_'.join(folders)}_{filename.split('.')[0]}_{timestamp}.log")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(formatter)
    logger.addHandler(fh)
    return logger

def extract_base_url(url):
	url_parts = urlparse(url)
	return f"{url_parts.scheme}://{url_parts.netloc}"


def get_accelify_user(excel_sheet):
    username, password = "SFTDVTester", "ht2jGMM2GnC3bwX7"
    if excel_sheet["B3"].value and excel_sheet["C3"].value: 
        username = excel_sheet["B3"].value
        password = excel_sheet["C3"].value
    return (username, password)

def click_distribute_button(driver):
    if "planng" in driver.current_url:
        buttons = driver.find_elements(By.TAG_NAME, "button")
        for button in buttons:
            if "Distribute" in button.text:
                driver.execute_script("arguments[0].click();", button)
                return
    else:
        driver.execute_script("$('#btnDistribute').click()")


def wait_dm(driver):
    wait = WebDriverWait(driver, 30)
    if "planng" in driver.current_url:
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "accelify-distribution-manager")))
    else:
        wait.until(
            EC.all_of(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#pnlEventFormsPackages table tr")),
                EC.presence_of_element_located((By.CSS_SELECTOR, "#pnlDistributeTo table tr")),
            )
        )
    time.sleep(5)


def get_packages_count(driver):
    if "planng" in driver.current_url:
        return len(driver.find_elements(By.CSS_SELECTOR, "accelify-packages table tr:not(.k-grid-norecords)")) - 1
    else:
        return driver.execute_script("return $('#pnlEventFormsPackages table tr:gt(0)').length")
    
    
def get_other_forms_count(driver):
    if "planng" in driver.current_url:
        return len(driver.find_elements(By.CSS_SELECTOR, "accelify-other-forms table tr:not(.k-grid-norecords)")) - 1
    else:
        return driver.execute_script("return $('#pnlEventOtherForms table tr:gt(0)').length")


def select_recepient(driver):
    if "planng" in driver.current_url:
        table = driver.find_element(By.CSS_SELECTOR, "#pnlDistributeTo table")
        rows = table.find_elements(By.CSS_SELECTOR, "tr:not(.k-grid-norecords)")
        if rows and len(rows) <= 1: return False
        if rows and len(rows) > 1: rows = rows[1:]
        for row in rows:
            cols = row.find_elements(By.CSS_SELECTOR, "td")
            dropdown = cols[-2].find_element(By.CSS_SELECTOR, "kendo-dropdownlist")
            dropdown.click()
            time.sleep(1)
            popup = driver.find_element(By.CSS_SELECTOR, "kendo-popup")
            options = popup.find_elements(By.CSS_SELECTOR, "ul>li")
            for option in options:
                if option.get_attribute('innerText') != "Collaboration Portal":
                    option.click()
                    break
            include_checkbox = cols[-1].find_element(By.TAG_NAME, "input")
            driver.execute_script("arguments[0].click();", include_checkbox)
            break
    else:
        driver.execute_script("""
            var first_elem = $("#pnlDistributeTo table tr:gt(0)").first();
            if (first_elem.length) {
                var lastColumnValue = first_elem.find("td:last").find("input[type='checkbox']").click();
                var dropdownlist = first_elem.find("td:last-of-type").prev().find("input[data-role='dropdownlist']").data("kendoDropDownList");
                for (var i = 0; i < dropdownlist.dataItems().length; ++i) {
                    if (dropdownlist.dataItems()[i].Name != "Collaboration Portal") {
                        dropdownlist.value(dropdownlist.dataItems()[i].LookupValueId);
                        dropdownlist.trigger("change");
                        break;
                    }
                }
            }
        """)
    return True


def select_package(driver, i):
    package_name = ""
    if "planng" in driver.current_url:
        packages = driver.find_elements(By.CSS_SELECTOR, "accelify-packages table tr")
        package = packages[i + 1].find_elements(By.TAG_NAME, "td")
        driver.execute_script("arguments[0].click();", package[1].find_element(By.TAG_NAME, "input"))
        package_name = package[2].get_attribute('innerText')
    else:
        packages = driver.find_elements(By.CSS_SELECTOR, "#pnlEventFormsPackages table tr")
        package = packages[i + 1].find_elements(By.CSS_SELECTOR, "td")
        driver.execute_script("arguments[0].click();", package[1].find_element(By.CSS_SELECTOR, "div > input"))
        package_name = package[2].get_attribute('innerText')
    time.sleep(1)
    return package_name

def select_other_form(driver, i):
    other_form_name = ""
    if "planng" in driver.current_url:
        other_forms = driver.find_elements(By.CSS_SELECTOR, "accelify-other-forms table tr")
        other_form = other_forms[i + 1].find_elements(By.TAG_NAME, "td")
        driver.execute_script("arguments[0].click();", other_form[1].find_element(By.TAG_NAME, "input"))
        other_form_name = other_form[2].get_attribute('innerText')
    else:
        other_forms = driver.find_elements(By.CSS_SELECTOR, "#pnlEventOtherForms table tr")
        other_form = other_forms[i + 1].find_elements(By.CSS_SELECTOR, "td")
        driver.execute_script("arguments[0].click();", other_form[0].find_element(By.CSS_SELECTOR, "div > input"))
        other_form_name = other_form[1].get_attribute('innerText')
    time.sleep(1)
    return other_form_name

    
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input_file", type=str)
    my_namespace = parser.parse_args()
    input_file = my_namespace.input_file

    logger = configure_logger("script-log", input_file)
    wb = load_workbook(input_file, data_only=True)
    wb_sheet = wb.active

    options = Options()
    options.add_argument('--disable-cache')
    options.add_argument('--disable-features=NetworkService')
    options.add_argument('--disable-session-storage')
    driver = webdriver.Chrome(options=options)

    username, password = get_accelify_user(wb_sheet)

    prev_base_url = ""
    base_url = ""
    for row in wb_sheet.iter_rows(min_row=5):
        url = row[1].value
        if not url or not validators.url(url): continue
        
        base_url = extract_base_url(url)
        if base_url != prev_base_url:
             SeleniumHelper.login_user(base_url, driver, username, password)
        prev_base_url = base_url

        prev_tab = driver.window_handles[0]
        driver.execute_script("window.open('');")
        driver.switch_to.window(prev_tab)
        curr_tab = driver.window_handles[1]
        driver.close()
        driver.switch_to.window(curr_tab)

        logger.info(f"Processing: {url}")
        driver.get(url)
        try:
            wait_dm(driver)
            packages_count = get_packages_count(driver)
            for i in range(packages_count):
                wait_dm(driver)
                if not select_recepient(driver): break
                package_name = select_package(driver, i)
                logger.info(f"      Processing: {package_name}")
                click_distribute_button(driver)
                time.sleep(5)
                driver.get(url)
            
            wait_dm(driver)
            other_forms_count = get_other_forms_count(driver)
            for i in range(other_forms_count):
                wait_dm(driver)
                if not select_recepient(driver): break
                package_name = select_other_form(driver, i)
                logger.info(f"      Processing other form: {package_name}")
                click_distribute_button(driver)
                time.sleep(5)
                driver.get(url)
        except Exception as ex:
            logger.exception(ex)

if __name__ == "__main__":
    main()