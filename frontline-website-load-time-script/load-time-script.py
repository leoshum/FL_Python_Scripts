import os
import sys
import time
import numpy as np
import validators
import argparse
import speedtest
import logging
from collections import namedtuple
from datetime import datetime
from urllib.parse import urlparse 
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException, NoSuchElementException, StaleElementReferenceException

package_path = os.path.abspath('..')
sys.path.append(package_path)
from frontline_selenium.selenium_helper import SeleniumHelper
from frontline_selenium.page_filler import PageFormFiller
from frontline_selenium.support_tech_helper import SupportTech

class HideBacktraceFormatter(logging.Formatter):
    def formatException(self, exc_info):
        tb = super().formatException(exc_info)
        return HideBacktraceFormatter.removeBackTrace(tb)
    
    def format(self, record: logging.LogRecord):
        record.msg = HideBacktraceFormatter.removeBackTrace(record.msg)
        return super().format(record)
    
    @staticmethod
    def removeBackTrace(record):
        tb_lines = str(record).splitlines()
        filtered_tb_lines = []
        skip_slce = False
        for line in tb_lines:
            if "Backtrace:" in line:
                skip_slce = True
            if "Traceback (most recent call last):" in line:
                skip_slce = False
            if not skip_slce:
                filtered_tb_lines.append(line)
        return "\n".join(filtered_tb_lines)
    
    
def split_path(path):
    folders = []
    head, tail = os.path.split(path)
    
    while tail:
        folders.insert(0, tail)
        head, tail = os.path.split(head)
    return folders


def is_excel_file_opened(filename):
    try:
        wb = load_workbook(filename)
        wb.save(filename)
        return False
    except:
        return True


def extract_base_url(url):
    url_parts = urlparse(url)
    return f"{url_parts.scheme}://{url_parts.netloc}"


def mark_form_as_invalid(row, color="FF0000"):
    row[0].font = Font(color=color)
    row[1].font = Font(color=color)


def measure_network_speed():
    try:
        st = speedtest.Speedtest()
        return round(st.download() / 1000000, 2)
    except:
        return -1


def specify_sheet_layout(sheet):
    sheet.move_range("N1:N1", rows=0, cols=9)
    sheet.move_range("E1:E1", rows=0, cols=9)
    sheet.move_range("N2:N2", rows=0, cols=9)
    sheet.move_range("E2:E2", rows=0, cols=9)


def reset_styles(cells):
    for cell in cells:
        cell.style = "Normal"


def flag_high_load_time(cells, threshold):
    for cell in cells:
        if cell.value != None and cell.value != "" and float(cell.value) > threshold:
            cell.font = Font(color="FF0000")


def measure_load_time(driver, url, loops, scenario):
    driver.get(url)
    measure_result = namedtuple("MeasureResult", ["first_measure", "min", "max", "mean"])
    totals = np.zeros(loops)
    is_first_measure = True
    first_measure = 0
    for j in range(loops):
        measured_time = scenario(driver)
        totals[j] = measured_time
        if is_first_measure:
            is_first_measure = False
            first_measure = measured_time
    return measure_result(first_measure, np.min(totals), np.max(totals), np.mean(totals))


def compare_measures(curr_cell, prev_cell, diff_cell):
    if curr_cell.value == None or prev_cell.value == None:
        reset_styles([curr_cell, prev_cell, diff_cell])
        return

    prev_row_float = 0.0
    try:
        prev_row_float = float(prev_cell.value[:prev_cell.value.index("(")])
    except:
        prev_row_float = float(prev_cell.value)

    try:
        diff = float(curr_cell.value) / prev_row_float
    except:
        diff = 0
    growth = (diff * 100) - 100
    diff_cell.value = f"{abs(growth):.2f}%"
    if growth < -10:
        diff_cell.fill = PatternFill(start_color="00FF00", fill_type = "solid")
    elif growth <= 10 and growth >= -10:
        diff_cell.fill = PatternFill(start_color="FFFF00", fill_type = "solid")
    else:
        diff_cell.fill = PatternFill(start_color="FF0000", fill_type = "solid")


def configure_logger(file_name: str, processing_filename: str) -> logging.Logger:
    logger = logging.getLogger("main")
    logger.setLevel(logging.DEBUG)

    formatter = HideBacktraceFormatter("%(asctime)s - %(message)s", datefmt="%m-%d-%y_%H:%M")
    timestamp = datetime.now().strftime("%m-%d-%y_%H-%M")
    parts = split_path(processing_filename)
    folders = parts[0:len(parts)-1]
    filename = parts[-1]
    fh = logging.FileHandler(f"{file_name}_{'_'.join(folders)}_{filename.split('.')[0]}_{timestamp}.log")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(formatter)
    logger.addHandler(fh)
    SeleniumHelper.setup_logger(logger)
    PageFormFiller.setup_logger(logger)
    return logger


def main():
    timestamp = datetime.now().strftime("%m-%d-%y_%H-%M")
    start_time = time.time()

    parser = argparse.ArgumentParser()
    parser.add_argument("input_file", type=str)
    parser.add_argument("--loops", type=int, default=3)
    parser.add_argument("--disable_save", action="store_true")
    parser.add_argument("--disable_filler", action="store_true", default=False)
    parser.add_argument("--idm_auth", action="store_true", default=False)
    my_namespace = parser.parse_args()

    input_file = my_namespace.input_file
    loops = my_namespace.loops
    disable_save = my_namespace.disable_save
    disable_filler = my_namespace.disable_filler
    idm_auth = my_namespace.idm_auth
    threshold = 6
    timeout = 30

    logger = configure_logger("script-log", input_file)
    SeleniumHelper.set_options({
        "disable_filler": disable_filler
    })
    print(f"Input file: {input_file}")
    if not os.path.isfile(input_file):
        print("Input file doesn't exist.")
        return
    if is_excel_file_opened(input_file):
        print(f"Close opened {input_file} file!")
        return
    
    network_speed = measure_network_speed()

    wb = load_workbook(input_file, data_only=True)
    wb_sheet = wb.active
    wb_sheet.cell(row=1, column=31).value = "."
    wb_sheet.cell(row=1, column=31).value = ""
    specify_sheet_layout(wb_sheet)

    options = Options()
    options.add_argument('--disable-cache')
    options.add_argument('--disable-features=NetworkService')
    options.add_argument('--disable-session-storage')
    #options.headless = True
    driver = webdriver.Chrome(options=options)
    head_cell_top = wb_sheet["F1"]
    head_cell_top.alignment = Alignment(horizontal='center')
    head_cell_bottom = wb_sheet["F2"]
    head_cell_bottom.alignment = Alignment(horizontal='center')

    for row in wb_sheet.iter_rows(min_row=5):
        for i in range(14, 18):
            row[i + 9].value = row[i].value

        for i in range(19, 22):
            row[i + 8].value = row[i].value

        reset_styles([row[14], row[15], row[16], 
                      row[17], row[23], row[24], 
                      row[25], row[26], row[27],
                      row[19], row[20], row[21],
                      row[9], row[13], row[18],
                      row[28], row[29], row[22]])

        for i in range(5, 9):
            row[i + 9].value = row[i].value

        for i in range(10, 13):
            row[i + 9].value = row[i].value

        flag_high_load_time([row[14], row[15], row[16], 
                             row[17], row[24], row[24], 
                             row[25], row[26], row[27],
                             row[19], row[20], row[21],
                             row[28], row[29]], threshold)
        for i in range(3, 13):
            row[i].value = ""
        reset_styles([row[5], row[6], row[7], row[8], row[10], row[11], row[12]])

    build_version = ""
    prev_base_url = ""

    base_url = ""
    is_first_row = True
    for row in wb_sheet.iter_rows(min_row=5):
        url = row[1].value
        if url == None or not validators.url(url):
            continue
        row[4].value = datetime.now().strftime('%y-%m-%d %H:%M:%S')
        print(f"{datetime.now().strftime('%y-%m-%d %H:%M:%S')}    {url}")
        logger.info(f"Processing: {url}")
        base_url = extract_base_url(url)
        if prev_base_url != base_url or is_first_row:
            if idm_auth:
                SupportTech.login(driver)
                time.sleep(3)
                SupportTech.open_website(driver, base_url, "SFTDVTester")
                driver.get(url)
            else:
                SeleniumHelper.login_user(base_url, driver, "SFTDVTester", "ht2jGMM2GnC3bwX7")
            build_version = SeleniumHelper.get_build_version(driver)
            is_first_row = False

        prev_tab = driver.window_handles[0]
        driver.execute_script("window.open('');")
        driver.switch_to.window(prev_tab)
        curr_tab = driver.window_handles[1]
        driver.close()
        driver.switch_to.window(curr_tab)

        scenario = SeleniumHelper.measure_form_page_load_time
        is_form_page_url = SeleniumHelper.is_form_page_url(url)
        if not is_form_page_url:
                scenario = SeleniumHelper.measure_standard_page_load_time

        try:
            (first_load_time, min_time, max_time, mean_time) = measure_load_time(driver, url, loops, scenario)
            reset_styles([row[0], row[1]])
        except Exception as e:
            (first_load_time, min_time, max_time, mean_time) = (timeout, timeout, timeout, timeout)
            row[3].value = f"Page speed measurement timeout"
            mark_form_as_invalid(row)


        error_in_page_loading = False
        try:
            page_title = driver.find_element(By.CSS_SELECTOR, "h1.page-title").text.strip()
            if "Error" in page_title or page_title == "Access Restricted":
                mark_form_as_invalid(row, color="0000FF")
                logger.debug(f"Error detected '{page_title}' in {url}")
                row[3].value = page_title
                error_in_page_loading = True
        except Exception as ex:
            logger.exception(ex)

        error_in_save = False
        if is_form_page_url and not disable_save and not error_in_page_loading:
            try:
                (first_save_time, min_save_time, max_save_time, mean_save_time) = measure_load_time(driver, url, loops, SeleniumHelper.measure_form_save_time)
            except ValueError as ex:
                mark_form_as_invalid(row)
                error_in_save = True
                logger.exception(ex)
                row[3].value = "Exception occured while saving the form!"
            except TimeoutException as ex:
                mark_form_as_invalid(row)
                error_in_save = True
                logger.exception(ex)
                row[3].value = "Saving timeout"
            except ElementClickInterceptedException as ex:
                mark_form_as_invalid(row, color="9933FF")
                error_in_save = True
                logger.exception(ex)
            except NoSuchElementException as ex:
                error_in_save = True
                logger.exception(ex)
            except StaleElementReferenceException as ex:
                mark_form_as_invalid(row, color="550000")
                error_in_save = True
                logger.exception(ex)

        row[5].value = f"{first_load_time:.2f}"
        row[6].value = f"{min_time:.2f}"
        row[7].value = f"{max_time:.2f}"
        row[8].value = f"{mean_time:.2f}"

        compare_measures(row[17], row[26], row[18])
        compare_measures(row[8], row[17], row[9])

        if is_form_page_url and not disable_save and not error_in_page_loading:
            if not error_in_save:
                row[10].value = f"{min_save_time:.2f}"
                row[11].value = f"{max_save_time:.2f}"
                row[12].value = f"{mean_save_time:.2f}"
                compare_measures(row[21], row[29], row[22])
                compare_measures(row[12], row[21], row[13])
            else:
                row[10].value = ""
                row[11].value = ""
                row[12].value = ""

        if disable_save or error_in_page_loading:
            row[10].value = ""
            row[11].value = ""
            row[12].value = ""

        flag_high_load_time([row[5], row[6], row[7], row[8], row[10], row[11], row[12]], threshold)
        prev_base_url = base_url
        head_cell_top.value = f"{build_version} {timestamp}"
        head_cell_bottom.value = f"{((time.time() - start_time) / 60):.2f}m, {network_speed}mb/s, loops: {loops}"
        wb.save(input_file)

if __name__ == "__main__":
    main()