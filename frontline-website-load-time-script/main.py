import time
import numpy as np
import validators
import argparse
import speedtest  
from collections import namedtuple
from datetime import datetime
from urllib.parse import urlparse 
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors, Font
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

def extract_base_url(url):
	url_parts = urlparse(url)
	return f"{url_parts.scheme}://{url_parts.netloc}"

def mark_form_as_invalid(row):
	row[0].font = Font(color="FF0000")
	row[1].font = Font(color="FF0000")

def measure_network_speed():
	st = speedtest.Speedtest()
	return st.download() / 1000000

def specify_sheet_layout(sheet):
	sheet.move_range("G1:G1", rows=0, cols=4)
	sheet.move_range("C1:C1", rows=0, cols=4)
	sheet.move_range("G2:G2", rows=0, cols=4)
	sheet.move_range("C2:C2", rows=0, cols=4)

def reset_styles(cells):
	for cell in cells:
		cell.style = "Normal"

def flag_high_load_time(cells, threshold):
	for cell in cells:
		if cell.value != None and float(cell.value) > threshold:
			cell.font = Font(color="FF0000")

def login_user(driver, url):
	driver.get(url)
	if "AcceliTrack" in driver.current_url:
		return
	username_field = driver.find_element("id", "UserName")
	password_field = driver.find_element("id", "Password")
	submit_btn = driver.find_element("id", "lnkLogin")
	username_field.send_keys("PMGMT")
	password_field.send_keys("8Huds(3d")
	submit_btn.click()

def measure_load_time(driver, url, timeout, loops):
	measure_result = namedtuple("MeasureResult", ["first_measure", "min", "max", "mean"])
	totals = np.zeros(loops)
	is_first_measure = True
	first_measure = 0
	for j in range(loops):
		start_time = time.time()
		driver.get(url)
		WebDriverWait(driver, timeout).until(
			EC.any_of(
				EC.presence_of_element_located((By.ID, "pnlForm")),
				EC.presence_of_element_located((By.ID, "pnlEventContent")),
				EC.presence_of_element_located((By.TAG_NAME, "accelify-forms-details")),
				EC.presence_of_element_located((By.TAG_NAME, "accelify-event-eligiblity-determination")),
				EC.presence_of_element_located((By.TAG_NAME, "accelify-progress-report"))
			)
		)
		total = time.time() - start_time
		totals[j] = total
		if is_first_measure:
			is_first_measure = False
			first_measure = total
	return measure_result(first_measure, np.min(totals), np.max(totals), np.mean(totals))

def compare_measures(curr_cell, prev_cell, diff_cell):
	if curr_cell.value == None or prev_cell.value == None:
		reset_styles([curr_cell, prev_cell, diff_cell])
		return

	prev_row_float = 0.0
	try:
		prev_row_float = float(prev_cell.value[:prev_cell.value.index("(")])
	except:
		prev_row_float = float(prev_cell.value)

	try:
		diff = float(curr_cell.value) / prev_row_float
	except:
		diff = 0
	growth = (diff * 100) - 100
	diff_cell.value = f"{abs(growth):.2f}%"
	if growth < -10:
		diff_cell.fill = PatternFill(start_color="00FF00", fill_type = "solid")
	elif growth <= 10 and growth >= -10:
		diff_cell.fill = PatternFill(start_color="FFFF00", fill_type = "solid")
	else:
		diff_cell.fill = PatternFill(start_color="FF0000", fill_type = "solid")

def main():
	timestamp = datetime.now().strftime("%m-%d-%y_%H-%M")
	start_time = time.time()
	parser = argparse.ArgumentParser()
	parser.add_argument("input_file", type=str)
	parser.add_argument("--loops", type=int, default=3)
	parser.add_argument("--enable_save", type=bool, default=True)
	my_namespace = parser.parse_args()

	network_speed = measure_network_speed()
	input_file = my_namespace.input_file
	loops = my_namespace.loops
	threshold = 6
	timeout = 30

	wb = load_workbook(input_file, data_only=True)
	wb_sheet = wb.active
	specify_sheet_layout(wb_sheet)
	
	driver = webdriver.Chrome()
	#driver.execute_cdp_cmd("Network.setCacheDisabled", {"cacheDisabled":True})
	options = Options()
	options.headless = True

	driver = webdriver.Chrome(options=options)

	head_cell_top = wb_sheet["C1"]
	head_cell_top.alignment = Alignment(horizontal='center')
	head_cell_bottom = wb_sheet["C2"]
	head_cell_bottom.alignment = Alignment(horizontal='center')

	build_version = ""
	prev_base_url = ""
	base_url = ""
	is_first_row = True
	for row in wb_sheet.iter_rows(min_row=5):
		form_url = row[1].value
		if form_url != None and validators.url(form_url):
			base_url = extract_base_url(form_url)
			if prev_base_url != base_url or is_first_row:
				login_user(driver, base_url)
				build_version = driver.find_element(By.CSS_SELECTOR, "span.version").text.replace("Version ", "")
				is_first_row = False
			row[21].value = row[12].value
			row[22].value = row[13].value
			row[23].value = row[14].value
			row[24].value = row[15].value

			row[25].value = row[17].value
			row[26].value = row[18].value
			row[27].value = row[19].value

			reset_styles([row[12], row[13], row[14], 
		 				  row[15], row[21], row[22], 
						  row[23], row[24], row[25],
						  row[17], row[18], row[19],
						  row[7], row[11], row[16],
						  row[26], row[27], row[20]])

			row[12].value = row[3].value
			row[13].value = row[4].value
			row[14].value = row[5].value
			row[15].value = row[6].value

			row[17].value = row[8].value
			row[18].value = row[9].value
			row[19].value = row[10].value

			flag_high_load_time([row[12], row[13], row[14], 
		 				  		 row[15], row[21], row[22], 
						  		 row[23], row[24], row[25],
						  		 row[17], row[18], row[19],
						  		 row[26], row[27]], threshold)

			try:
				(first_load_time, min_time, max_time, mean_time) = measure_load_time(driver, form_url, timeout, loops)
				reset_styles([row[0], row[1]])
			except TimeoutException:
				(first_load_time, min_time, max_time, mean_time) = (timeout, timeout, timeout, timeout)
				mark_form_as_invalid(row)

			row[3].value = f"{first_load_time:.2f}"
			row[4].value = f"{min_time:.2f}"
			row[5].value = f"{max_time:.2f}"
			row[6].value = f"{mean_time:.2f}"

			compare_measures(row[15], row[24], row[16])
			compare_measures(row[6], row[15], row[7])
			reset_styles([row[3], row[4], row[5], row[6]])
			flag_high_load_time([row[3], row[4], row[5], row[6]], threshold)
			prev_base_url = base_url
	head_cell_top.value = f"{build_version} {timestamp}"
	head_cell_bottom.value = f"{((time.time() - start_time) / 60):.2f}m, {network_speed:.2f}mb/s, loops: {loops}"
	wb.save(input_file)

if __name__ == "__main__":
	main()