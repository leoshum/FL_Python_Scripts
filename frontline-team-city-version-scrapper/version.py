from asyncio import get_event_loop, sleep
from datetime import datetime
from json import load
from logging import INFO, basicConfig, getLogger
from os import environ, path

import openpyxl
from aiohttp import ClientSession, TCPConnector, client_exceptions
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from packaging import version


class VersionScrapper:
    def __init__(self, domain) -> None:
        self.session = None
        self.base_url = f"https://{domain}/app/rest"

        file_name = path.splitext(path.basename(__file__))[0]
        self.root_directory = path.dirname(__file__)

        self.version_path = f'{self.root_directory}\\versions.xlsx'

        with open(f'{self.root_directory}\\projects.json',mode='r',encoding='UTF-8') as configuration:
            self.configuration = load(configuration)

        self.logger = getLogger(file_name)
        basicConfig(level=INFO)

        token_name = 'TEAM_CITY_TOKEN'
        self.token = environ.get(token_name)
        if self.token is None:
            self.logger.error(msg=f"TeamCity bearer token not found in environments variable with name [{token_name}]!")
            exit()
        pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        return

    async def get(self, endpoint, session):
        headers = {
        "Authorization" : f"Bearer {self.token}",
        "Accept" : "application/json"}

        attempt = 0
        while attempt < 3:
            try:
                attempt += 1
                async with session.get(f'{self.base_url}/{endpoint}', headers=headers) as resp:
                    if resp.status != 200:
                        raise Exception("Error in API call: " + await resp.text())
                    return await resp.json()
            except client_exceptions.ClientOSError as e:
                print(f'Error making request: {e}')
                await sleep(1)
            except client_exceptions.ServerDisconnectedError as e:
                print(f'Server disconnected error: {e}')
                await sleep(1)
            except Exception as e:
                print(f'Unexpected error occurred: {e}')
                await sleep(1)  

    async def projects(self, parent, session):
        url = f'projects/id:{parent}'
        return await self.get(url, session)
        
    async def builds(self, project, session):
        url = f'builds?locator=defaultFilter:false,buildType:{project},status:SUCCESS'#start:{start},
        return await self.get(url, session)
    
    async def build(self, build, session, fields = ['startDate']):
        url = f'builds/id:{build}?fields={",".join(fields)}'
        return await self.get(url, session)

    async def process_builds(self, builds, session):
        result = []
        major_proccessing = 0
        last_major = version.Version('0.0')

        for build in builds.get('build'):
            builds_version = version.parse(build.get('number'))
            if builds_version.minor != last_major.minor:
                last_major = builds_version
                major_proccessing += 1
                if major_proccessing > 3:
                    break
                major = {
                    'version' : '.'.join([str(builds_version.major), str(builds_version.minor), str(builds_version.micro)]),
                    'minors' : []
                }
                result.append(major)
            details = await self.build(build.get('id'), session)
            major.get('minors').append({
                'number' : builds_version.release[3],
                'time' : datetime.strptime(details.get('startDate'), '%Y%m%dT%H%M%S%z')
            })
        return result

    async def start(self):
        result_clients = []

        async with ClientSession(connector=TCPConnector(verify_ssl=False),trust_env=True) as session:
            root_projects = await self.projects('DeployNewArchitecture', session)
            for cfg_environment in self.configuration:
                for environment in root_projects.get('projects').get('project'):

                    if environment.get('name') == cfg_environment.get('Environment'):

                        client_projects = await self.projects(environment.get('id'), session)
                        for client in client_projects.get('buildTypes').get('buildType'):
                            for cfg_client in cfg_environment.get('Clients'):

                                if client.get('name') == cfg_client:
                                    builds = await self.builds(client.get('id'), session)                                    
                                    result_clients.append({
                                        "name" : cfg_client,
                                        'majors' : await self.process_builds(builds, session)
                                    })
        self.save(result_clients)

    
    def updateLength(self, value, width, name):
        length = len(value)
        if length > width[name]:
            width[name] = length

    def save(self, clients):
        width = {
            'Name' : 0,
            'Time' : 0
        }
        bold_font = Font(bold=True)
        
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Versions"
        current_row = 1
        for client in clients:
            minors_length = 0
            sheet.append([client.get('name')])
            sheet.cell(row=sheet.max_row, column=1).font = bold_font

            # sheet.append([])
            self.updateLength(client.get('name'), width, 'Name')

            majors = client.get('majors')
            for major in majors:
                sheet.append([major.get('version')])
                sheet.cell(row=sheet.max_row, column=1).font = bold_font
                self.updateLength(major.get('version'), width, 'Name')

                minors = major.get('minors')            
                for minor in minors:
                    time = minor.get('time').strftime('%Y-%m-%d')
                    sheet.append([minor.get('number'), time])
                    self.updateLength(str(minor.get('number')), width, 'Name')
                    self.updateLength(time, width, 'Time')
                    minors_length += 1
            length = len(majors) + minors_length + 1
            # sheet.row_dimensions.group(current_row + 1, current_row + length, hidden=False)
            current_row += length + 1

        
        for i, column_width in enumerate(width.values(),1):  # ,1 to start at 1
            sheet.column_dimensions[get_column_letter(i)].width = column_width + 3
        book.save(filename=self.version_path)

with VersionScrapper('teams.acceliplan.com') as scrapper:
    get_event_loop().run_until_complete(scrapper.start())